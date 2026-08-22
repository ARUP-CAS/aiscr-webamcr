"""Testy pro trvalé ukládání importního XLSX reportu na disk.

Zákazník na schůzce potvrdil, že durabilita reportu se řeší uložením XLSX do souborového
adresáře (ne DB tabulkou) — testy zde pokrývají sdílenou kontrolu adresáře
(``core.utils.check_import_report_directory``) a mechanismus ukládání/sestavení reportu
(``cron.tasks.get_or_create_import_report_path`` / ``build_import_report_dataframe`` /
``save_import_report_to_disk``).
"""

import json
import os
import tempfile

import openpyxl
from core.setting_models import CustomAdminSettings
from core.tests.fake_redis import FakeRedis
from core.utils import (
    ImportReportIndexError,
    check_import_report_directory,
    check_import_report_index_files_exist,
    read_import_report_index,
    upsert_import_report_index_entry,
)
from cron.tasks import (
    build_import_fedora_target_dataframe,
    build_import_report_dataframe,
    get_or_create_import_report_path,
    save_import_report_to_disk,
)
from django.test import TestCase

JOB_ID = "report-test-job"


def _configure_import_directory(directory_path):
    """Uloží ``CustomAdminSettings`` s ``DIRECTORY_PATH`` ukazujícím na daný adresář."""
    CustomAdminSettings.objects.create(
        item_group="import",
        item_id="import_directory_settings",
        value=json.dumps({"DIRECTORY_PATH": directory_path}),
    )


class CheckImportReportDirectoryTest(TestCase):
    """Testy pro ``core.utils.check_import_report_directory``."""

    def test_missing_setting_returns_error(self):
        """Bez uloženého nastavení vrátí chybu a obě cesty ``None``."""
        directory_path, reports_path, error = check_import_report_directory()
        self.assertIsNone(directory_path)
        self.assertIsNone(reports_path)
        self.assertIsNotNone(error)

    def test_missing_directory_path_key_returns_error(self):
        """Nastavení bez klíče ``DIRECTORY_PATH`` je považováno za nekonfigurované."""
        CustomAdminSettings.objects.create(
            item_group="import", item_id="import_directory_settings", value=json.dumps({})
        )
        directory_path, reports_path, error = check_import_report_directory()
        self.assertIsNone(directory_path)
        self.assertIsNotNone(error)

    def test_nonexistent_directory_returns_error(self):
        """Nakonfigurovaný, ale neexistující adresář musí vrátit chybu."""
        _configure_import_directory("/nonexistent/path/for/webamcr/import/tests")
        directory_path, reports_path, error = check_import_report_directory()
        self.assertIsNone(directory_path)
        self.assertIsNotNone(error)

    def test_valid_directory_creates_reports_subdirectory(self):
        """Platný adresář vede k vytvoření podadresáře ``reports``, pokud ještě neexistuje."""
        with tempfile.TemporaryDirectory() as tmp:
            _configure_import_directory(tmp)
            directory_path, reports_path, error = check_import_report_directory()
            self.assertIsNone(error)
            self.assertEqual(directory_path, tmp)
            self.assertEqual(reports_path, os.path.join(tmp, "reports"))
            self.assertTrue(os.path.isdir(reports_path))

    def test_existing_reports_subdirectory_is_reused(self):
        """Již existující podadresář ``reports`` se nepřepisuje ani nehlásí jako chyba."""
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            marker = os.path.join(reports_dir, "marker.txt")
            with open(marker, "w") as fh:
                fh.write("existing")
            _configure_import_directory(tmp)
            directory_path, reports_path, error = check_import_report_directory()
            self.assertIsNone(error)
            self.assertTrue(os.path.isfile(marker), "Existující obsah adresáře reports nesmí být smazán.")

    def test_unwritable_reports_directory_returns_error(self):
        """Adresář ``reports`` bez práva zápisu musí selhat s chybou (ne tichý no-op)."""
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            os.chmod(reports_dir, 0o500)
            try:
                _configure_import_directory(tmp)
                directory_path, reports_path, error = check_import_report_directory()
                self.assertIsNone(directory_path)
                self.assertIsNotNone(error)
            finally:
                os.chmod(reports_dir, 0o700)

    def test_check_writable_false_skips_writability_probe(self):
        """``check_writable=False`` (re-validace uvnitř běžícího importu) neprovádí zápis."""
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            os.chmod(reports_dir, 0o500)
            try:
                _configure_import_directory(tmp)
                directory_path, reports_path, error = check_import_report_directory(check_writable=False)
                self.assertIsNone(error)
                self.assertEqual(directory_path, tmp)
            finally:
                os.chmod(reports_dir, 0o700)


def _populate_report_redis(fake_redis, phase="importing"):
    """Naplní ``FakeRedis`` typickými daty jedné importní úlohy pro sestavení reportu."""
    fake_redis.set("import_data_phase_{}".format(JOB_ID), phase)
    fake_redis.set(
        "import_data_validation_results_{}".format(JOB_ID),
        json.dumps(
            [
                {"item_order": 0, "file_name": "dokument.csv", "primary_key_import": "C-1", "validation_result": "ok"},
                {"item_order": 1, "file_name": "dokument.csv", "primary_key_import": "C-2", "validation_result": "ok"},
            ]
        ),
    )
    fake_redis.set("import_data_primary_keys_{}".format(JOB_ID), json.dumps({"0": "1", "1": "2"}))
    fake_redis.rpush("import_data_progress_ids_{}".format(JOB_ID), 0)
    fake_redis.rpush("import_data_progress_ids_{}".format(JOB_ID), 1)
    fake_redis.rpush("import_data_progress_details_tr_{}".format(JOB_ID), "cron.tasks.run_data_import.success")
    fake_redis.rpush("import_data_progress_details_tr_{}".format(JOB_ID), "cron.tasks.run_data_import.success")
    fake_redis.set(
        "import_data_history_record_result_tr_{}".format(JOB_ID), json.dumps({"0": "hist-ok", "1": "hist-ok"})
    )
    fake_redis.set("import_fedora_result_tr_{}".format(JOB_ID), json.dumps({"0": ["fedora-uid (C-1)"]}))


class BuildImportReportDataframeTest(TestCase):
    """Testy pro ``cron.tasks.build_import_report_dataframe``.

    Regrese pro bug objevený při implementaci: volající uvnitř ``cron.tasks`` používají
    ``RedisConnector.get_connection()`` (bytový režim), na rozdíl od ``core.views``, které používá
    dekódující spojení — bez normalizace na ``str`` by fáze i per-záznamové sloupce v na disk
    ukládaném reportu zůstaly prázdné, přestože stažený report by je correctly zobrazoval.
    """

    def test_bytes_mode_connector_still_matches_per_record_columns(self):
        """Report sestavený nad bytovým Redis spojením musí mít stejný obsah jako nad dekódujícím."""
        fake_redis_bytes = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis_bytes)
        df_bytes, phase_bytes = build_import_report_dataframe(JOB_ID, fake_redis_bytes)

        fake_redis_decoded = FakeRedis(decode_responses=True)
        _populate_report_redis(fake_redis_decoded)
        df_decoded, phase_decoded = build_import_report_dataframe(JOB_ID, fake_redis_decoded)

        self.assertEqual(phase_bytes, "importing")
        self.assertEqual(phase_bytes, phase_decoded)
        self.assertEqual(df_bytes.shape, df_decoded.shape)
        # Compare stringified contents rather than translated column headers (locale-independent).
        self.assertEqual(df_bytes.astype(str).values.tolist(), df_decoded.astype(str).values.tolist())
        # The partial-report banner row must be present for an in-progress phase in BOTH modes —
        # this is exactly the check that silently broke under bytes mode before the fix.
        self.assertEqual(len(df_bytes), 3)  # banner row + 2 data rows

    def test_finished_phase_has_no_partial_banner(self):
        """Report sestavený pro terminální fázi nesmí obsahovat banner částečného reportu."""
        fake_redis = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis, phase="finished")
        df, phase = build_import_report_dataframe(JOB_ID, fake_redis)
        self.assertEqual(phase, "finished")
        self.assertEqual(len(df), 2)


class SaveImportReportToDiskTest(TestCase):
    """Testy pro ``cron.tasks.save_import_report_to_disk`` a ``get_or_create_import_report_path``."""

    def test_path_is_stable_across_calls(self):
        """Opakované volání pro stejný ``job_id`` musí vracet stejnou cestu (validace i import píší
        do stejného souboru)."""
        fake_redis = FakeRedis(decode_responses=False)
        with tempfile.TemporaryDirectory() as tmp:
            first = get_or_create_import_report_path(JOB_ID, fake_redis, tmp)
            second = get_or_create_import_report_path(JOB_ID, fake_redis, tmp)
            self.assertEqual(first, second)
            self.assertTrue(first.startswith(tmp))
            self.assertIn(JOB_ID, first)
            self.assertTrue(first.endswith(".xlsx"))

    def test_save_writes_readable_xlsx_and_sets_redis_flag(self):
        """Úspěšné uložení zapíše čitelný XLSX (s listem ``Fedora``) a nastaví Redis příznak."""
        fake_redis = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis)
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            result_path = save_import_report_to_disk(JOB_ID, fake_redis, reports_dir)
            self.assertIsNotNone(result_path)
            self.assertTrue(os.path.isfile(result_path))
            workbook = openpyxl.load_workbook(result_path)
            self.assertIn("Import", workbook.sheetnames)
            self.assertGreater(workbook["Import"].max_row, 1)
            self.assertIn("Fedora", workbook.sheetnames)
            saved_flag = fake_redis.get("import_data_report_saved_path_{}".format(JOB_ID))
            self.assertEqual(saved_flag.decode("utf-8") if isinstance(saved_flag, bytes) else saved_flag, result_path)

    def test_save_upserts_report_index_entry(self):
        """Uložení reportu musí zapsat/aktualizovat záznam v JSON indexu adresáře reportů."""
        fake_redis = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis, phase="importing")
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            result_path = save_import_report_to_disk(JOB_ID, fake_redis, reports_dir)

            entries = read_import_report_index(reports_dir)
            self.assertEqual(len(entries), 1)
            self.assertEqual(entries[0]["job_id"], JOB_ID)
            self.assertEqual(entries[0]["file_name"], os.path.basename(result_path))
            self.assertEqual(entries[0]["stage"], "importing")

            # A later save for the same job_id updates the entry in place instead of appending.
            _populate_report_redis(fake_redis, phase="finished")
            save_import_report_to_disk(JOB_ID, fake_redis, reports_dir)
            entries = read_import_report_index(reports_dir)
            self.assertEqual(len(entries), 1)
            self.assertEqual(entries[0]["stage"], "finished")

    def test_second_save_overwrites_same_file_not_a_new_one(self):
        """Druhé volání musí přepsat stejný soubor (stejná cesta), ne vytvořit další."""
        fake_redis = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis)
        with tempfile.TemporaryDirectory() as tmp:
            reports_dir = os.path.join(tmp, "reports")
            os.makedirs(reports_dir)
            first_path = save_import_report_to_disk(JOB_ID, fake_redis, reports_dir)
            _populate_report_redis(fake_redis, phase="finished")
            second_path = save_import_report_to_disk(JOB_ID, fake_redis, reports_dir)
            self.assertEqual(first_path, second_path)
            # Only one XLSX must exist across both saves — the JSON index (index.json) is the one
            # other file the directory now legitimately holds.
            xlsx_files = [name for name in os.listdir(reports_dir) if name.endswith(".xlsx")]
            self.assertEqual(len(xlsx_files), 1)

    def test_write_failure_is_swallowed_and_logged(self):
        """Selhání zápisu (needostupný adresář) nesmí vyhodit výjimku — jen se zaloguje a vrátí None."""
        fake_redis = FakeRedis(decode_responses=False)
        _populate_report_redis(fake_redis)
        nonexistent_dir = "/nonexistent/reports/dir/for/webamcr/tests"
        result = save_import_report_to_disk(JOB_ID, fake_redis, nonexistent_dir)
        self.assertIsNone(result)
        self.assertIsNone(fake_redis.get("import_data_report_saved_path_{}".format(JOB_ID)))


class BuildImportFedoraTargetDataframeTest(TestCase):
    """Testy pro ``cron.tasks.build_import_fedora_target_dataframe`` (list ``Fedora``)."""

    def test_empty_targets_still_has_columns(self):
        """Bez žádného Fedora cíle musí DataFrame mít prázdné řádky, ale definované sloupce."""
        fake_redis = FakeRedis(decode_responses=False)
        df = build_import_fedora_target_dataframe(JOB_ID, fake_redis)
        self.assertEqual(len(df), 0)
        self.assertEqual(len(df.columns), 3)

    def test_translates_result_and_keeps_raw_transaction_data(self):
        """Sloupec výsledku se přeloží z translation ID, ident_cely a transaction_uid zůstanou syrová data."""
        fake_redis = FakeRedis(decode_responses=False)
        fake_redis.set(
            "import_fedora_target_results_tr_{}".format(JOB_ID),
            json.dumps(
                [
                    {
                        "ident_cely": "C-101-2024",
                        "transaction_uid": "uuid-123",
                        "result": "cron.tasks.run_data_import.fedora_target_success",
                    },
                    {
                        "ident_cely": None,
                        "transaction_uid": None,
                        "result": "cron.tasks.run_data_import.fedora_target_error",
                    },
                ]
            ),
        )
        df = build_import_fedora_target_dataframe(JOB_ID, fake_redis)
        self.assertEqual(len(df), 2)
        values = df.astype(str).values.tolist()
        self.assertEqual(values[0][0], "C-101-2024")
        self.assertEqual(values[0][1], "uuid-123")

    def test_records_without_a_fedora_target_get_a_skipped_row_with_blank_transaction(self):
        """Záznam bez Fedora cíle dostane placeholder ``fedora_target_skipped`` s prázdným transaction_uid."""
        fake_redis = FakeRedis(decode_responses=False)
        fake_redis.set(
            "import_data_validation_results_{}".format(JOB_ID),
            json.dumps(
                [
                    {"item_order": 0, "file_name": "x.csv", "primary_key_import": "C-1", "validation_result": "ok"},
                ]
            ),
        )
        fake_redis.set(
            "import_fedora_result_tr_{}".format(JOB_ID),
            json.dumps({"0": ["cron.tasks.run_data_import.fedora_skipped"]}),
        )
        df = build_import_fedora_target_dataframe(JOB_ID, fake_redis)
        self.assertEqual(len(df), 1)
        values = df.astype(str).values.tolist()
        self.assertEqual(values[0][0], "C-1")
        self.assertEqual(values[0][1], "")

    def test_waiting_on_files_is_not_reported_as_skipped(self):
        """Záznam čekající na fázi souborů (jiný placeholder) se nesmí objevit jako ``skipped``."""
        fake_redis = FakeRedis(decode_responses=False)
        fake_redis.set(
            "import_fedora_result_tr_{}".format(JOB_ID),
            json.dumps({"0": ["cron.tasks.run_data_import.fedora_waiting_data_import"]}),
        )
        df = build_import_fedora_target_dataframe(JOB_ID, fake_redis)
        self.assertEqual(len(df), 0)


class ImportReportIndexTest(TestCase):
    """Testy pro JSON index reportů (``core.utils.read_import_report_index`` a spol.)."""

    def test_missing_index_reads_as_empty_list(self):
        """Adresář bez ``index.json`` (report ještě nebyl uložen) se čte jako prázdný seznam."""
        with tempfile.TemporaryDirectory() as tmp:
            self.assertEqual(read_import_report_index(tmp), [])

    def test_upsert_adds_and_updates_entry(self):
        """Zápis pro nový ``job_id`` přidá záznam; zápis pro existující ``job_id`` ho nahradí."""
        with tempfile.TemporaryDirectory() as tmp:
            upsert_import_report_index_entry(tmp, "job-a", "a.xlsx", "importing")
            upsert_import_report_index_entry(tmp, "job-b", "b.xlsx", "importing")
            entries = read_import_report_index(tmp)
            self.assertEqual(len(entries), 2)

            upsert_import_report_index_entry(tmp, "job-a", "a.xlsx", "finished")
            entries = read_import_report_index(tmp)
            self.assertEqual(len(entries), 2)
            job_a = next(e for e in entries if e["job_id"] == "job-a")
            self.assertEqual(job_a["stage"], "finished")

    def test_corrupt_index_is_treated_as_empty(self):
        """Poškozený JSON index nesmí shodit čtenáře — chová se, jako by byl prázdný."""
        with tempfile.TemporaryDirectory() as tmp:
            with open(os.path.join(tmp, "index.json"), "w") as fh:
                fh.write("{not valid json")
            self.assertEqual(read_import_report_index(tmp), [])

    def test_check_files_exist_flags_missing_and_raises(self):
        """Chybějící soubor na disku se musí propsat do ``exists`` a vyvolat výjimku."""
        with tempfile.TemporaryDirectory() as tmp:
            with open(os.path.join(tmp, "present.xlsx"), "wb"):
                pass
            entries = [
                {"job_id": "present-job", "file_name": "present.xlsx"},
                {"job_id": "missing-job", "file_name": "missing.xlsx"},
            ]
            with self.assertRaises(ImportReportIndexError) as ctx:
                check_import_report_index_files_exist(entries, tmp)
            self.assertEqual(ctx.exception.missing_job_ids, ["missing-job"])
            # Entries are annotated in place even though the call raised.
            self.assertTrue(entries[0]["exists"])
            self.assertFalse(entries[1]["exists"])

    def test_check_files_exist_no_missing_does_not_raise(self):
        """Když všechny soubory existují, funkce nevyhodí výjimku."""
        with tempfile.TemporaryDirectory() as tmp:
            with open(os.path.join(tmp, "present.xlsx"), "wb"):
                pass
            entries = [{"job_id": "present-job", "file_name": "present.xlsx"}]
            check_import_report_index_files_exist(entries, tmp)
            self.assertTrue(entries[0]["exists"])
